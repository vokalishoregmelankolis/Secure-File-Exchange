import os
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'jpg', 'jpeg', 'png', 'gif', 'pdf'}

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def generate_file_id():
    """Generate a unique file identifier using UUID"""
    return str(uuid.uuid4())

def get_file_type(filename):
    """Determine file type category"""
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    if ext in ['xlsx', 'xls']:
        return 'spreadsheet'
    elif ext in ['jpg', 'jpeg', 'png', 'gif']:
        return 'image'
    elif ext == 'pdf':
        return 'pdf'
    else:
        return 'other'

def format_file_size(size_bytes):
    """Format file size in human-readable format"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f'{size_bytes:.2f} {unit}'
        size_bytes /= 1024.0
    return f'{size_bytes:.2f} TB'

def format_execution_time(time_seconds):
    """Format execution time in human-readable format"""
    if time_seconds < 0.001:
        return f'{time_seconds * 1000000:.2f} μs'
    elif time_seconds < 1:
        return f'{time_seconds * 1000:.2f} ms'
    else:
        return f'{time_seconds:.4f} s'

def create_financial_report_template(output_path):
    """
    Create an Excel template for financial reports
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Financial Report'
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'CONFIDENTIAL FINANCIAL REPORT'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    # Company Information
    ws['A3'] = 'Company Name:'
    ws['B3'] = '[Enter Company Name]'
    ws['A4'] = 'Report Period:'
    ws['B4'] = '[e.g., Q1 2024]'
    ws['A5'] = 'Department:'
    ws['B5'] = '[Enter Department]'
    
    # Revenue Section
    ws['A7'] = 'REVENUE SUMMARY'
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws.merge_cells('A7:D7')
    
    ws['A8'] = 'Total Revenue'
    ws['B8'] = 0
    ws['A9'] = 'Total Expenses'
    ws['B9'] = 0
    ws['A10'] = 'Net Profit'
    ws['B10'] = '=B8-B9'
    ws['B10'].font = Font(bold=True)
    
    # Budget Section
    ws['A12'] = 'BUDGET ANALYSIS'
    ws['A12'].font = header_font
    ws['A12'].fill = header_fill
    ws.merge_cells('A12:D12')
    
    ws['A13'] = 'Budget Allocated'
    ws['B13'] = 0
    ws['A14'] = 'Budget Spent'
    ws['B14'] = 0
    ws['A15'] = 'Variance'
    ws['B15'] = '=B13-B14'
    ws['B15'].font = Font(bold=True)
    
    # Notes Section
    ws['A17'] = 'NOTES'
    ws['A17'].font = header_font
    ws['A17'].fill = header_fill
    ws.merge_cells('A17:D17')
    
    ws['A18'] = '[Add any additional notes or comments here]'
    ws.merge_cells('A18:D20')
    
    # Apply borders
    for row in range(3, 16):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    wb.save(output_path)
    return output_path

def parse_financial_report(file_path):
    """
    Parse financial data from Excel file
    Returns: Dictionary of financial data
    """
    try:
        df = pd.read_excel(file_path, header=None)
        
        # Extract data from specific cells
        data = {
            'company_name': str(df.iloc[2, 1]) if pd.notna(df.iloc[2, 1]) else '',
            'report_period': str(df.iloc[3, 1]) if pd.notna(df.iloc[3, 1]) else '',
            'department': str(df.iloc[4, 1]) if pd.notna(df.iloc[4, 1]) else '',
            'total_revenue': str(df.iloc[7, 1]) if pd.notna(df.iloc[7, 1]) else '0',
            'total_expenses': str(df.iloc[8, 1]) if pd.notna(df.iloc[8, 1]) else '0',
            'net_profit': str(df.iloc[9, 1]) if pd.notna(df.iloc[9, 1]) else '0',
            'budget_allocated': str(df.iloc[12, 1]) if pd.notna(df.iloc[12, 1]) else '0',
            'budget_spent': str(df.iloc[13, 1]) if pd.notna(df.iloc[13, 1]) else '0',
            'variance': str(df.iloc[14, 1]) if pd.notna(df.iloc[14, 1]) else '0',
            'notes': str(df.iloc[17, 0]) if pd.notna(df.iloc[17, 0]) else ''
        }
        
        return data
    except Exception as e:
        print(f'Error parsing financial report: {e}')
        return None

def log_operation(user_id, file_id, operation, algorithm=None, success=True, error_message=None, ip_address=None):
    """
    Log encryption/decryption operations
    """
    from app.models import EncryptionLog
    from app import db
    
    log = EncryptionLog(
        user_id=user_id,
        file_id=file_id,
        operation=operation,
        algorithm=algorithm,
        success=success,
        error_message=error_message,
        ip_address=ip_address,
        timestamp=datetime.utcnow()
    )
    db.session.add(log)
    db.session.commit()


def log_crypto_operation(user_id, operation, details=None, success=True, error_message=None, ip_address=None):
    """
    Log cryptographic operations (key generation, wrapping, unwrapping, etc.)
    
    This function ensures that all cryptographic operations are logged for audit purposes
    while ensuring that no sensitive key material is ever included in the logs.
    
    Args:
        user_id: ID of the user performing the operation
        operation: Type of operation (keypair_generated, key_wrapped, key_unwrapped, 
                   private_key_decrypted, access_granted, access_revoked)
        details: Additional details about the operation (must not contain sensitive keys)
        success: Whether the operation succeeded
        error_message: Error message if operation failed (must not contain sensitive keys)
        ip_address: IP address of the user
    
    Operations:
        - keypair_generated: RSA key pair generation
        - key_wrapped: Symmetric key wrapped with RSA public key
        - key_unwrapped: Symmetric key unwrapped with RSA private key
        - private_key_decrypted: Private key decrypted with password
        - access_granted: Access request approved
        - access_revoked: Access revoked
    
    Security:
        This function sanitizes all inputs to ensure no plaintext keys, passwords,
        or other sensitive cryptographic material is logged.
    """
    from app.models import CryptoLog
    from app import db
    
    # Sanitize details and error_message to ensure no sensitive data is logged
    sanitized_details = _sanitize_log_data(details) if details else None
    sanitized_error = _sanitize_log_data(error_message) if error_message else None
    
    log = CryptoLog(
        user_id=user_id,
        operation=operation,
        details=sanitized_details,
        success=success,
        error_message=sanitized_error,
        ip_address=ip_address,
        timestamp=datetime.utcnow()
    )
    db.session.add(log)
    
    try:
        db.session.commit()
    except Exception as e:
        # If logging fails, rollback but don't raise - logging should not break operations
        db.session.rollback()
        print(f"Warning: Failed to log crypto operation: {e}")


def _sanitize_log_data(data):
    """
    Sanitize log data to remove any sensitive cryptographic material.
    
    This function ensures that plaintext keys, passwords, and other sensitive
    data are never logged. It looks for common patterns and removes them.
    
    Args:
        data: String data to sanitize
    
    Returns:
        Sanitized string with sensitive data removed
    """
    if not data or not isinstance(data, str):
        return data
    
    # List of sensitive keywords that should not appear in logs
    sensitive_keywords = [
        'private_key', 'private key', 'privatekey',
        'symmetric_key', 'symmetric key', 'symmetrickey',
        'password', 'passphrase', 'secret',
        'plaintext', 'plain text',
        'unwrapped_key', 'unwrapped key',
        'decrypted_key', 'decrypted key',
        'encryption_key', 'encryption key',
        'aes_key', 'aes key',
        'des_key', 'des key',
        'rc4_key', 'rc4 key'
    ]
    
    # Convert to lowercase for case-insensitive checking
    data_lower = data.lower()
    
    # Check if any sensitive keywords appear in the data
    for keyword in sensitive_keywords:
        if keyword in data_lower:
            # If sensitive keyword found, return a sanitized message
            return f"[REDACTED: Log contained sensitive keyword '{keyword}']"
    
    # Check for base64-encoded data (common for keys)
    # Base64 strings are typically long alphanumeric strings with +/= characters
    import re
    base64_pattern = r'[A-Za-z0-9+/]{40,}={0,2}'
    if re.search(base64_pattern, data):
        # Replace base64-like strings with placeholder
        data = re.sub(base64_pattern, '[REDACTED: Base64 data]', data)
    
    # Check for hex-encoded data (another common key format)
    hex_pattern = r'[0-9a-fA-F]{64,}'
    if re.search(hex_pattern, data):
        # Replace hex strings with placeholder
        data = re.sub(hex_pattern, '[REDACTED: Hex data]', data)
    
    # Limit length to prevent logging of large binary data
    max_length = 500
    if len(data) > max_length:
        data = data[:max_length] + '... [TRUNCATED]'
    
    return data