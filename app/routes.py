from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.models import User, EncryptedFile, SharedFile, FinancialReport, PerformanceMetric, AccessRequest, UserRole, CryptoLog
from app.forms import RegistrationForm, LoginForm, FileUploadForm, FinancialReportForm, ShareFileForm, AccessRequestForm, DecryptKeyForm
from app.crypto_utils import CryptoEngine, encrypt_financial_data, decrypt_financial_data
from app.utils import (allowed_file, generate_file_id, get_file_type, format_file_size, 
                       format_execution_time, create_financial_report_template, 
                       parse_financial_report, log_operation, log_crypto_operation)
import os
from io import BytesIO
from datetime import datetime
from sqlalchemy import or_
from urllib.parse import urlparse


def is_safe_url(target):
    """Check if URL is safe for redirect (local URL only).
    
    Prevents open redirect vulnerabilities by ensuring the target URL
    is either a relative path (no scheme/netloc) or matches our host exactly.
    """
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(target)
    
    # Relative paths (no scheme and no netloc) are always safe
    if test_url.scheme == '' and test_url.netloc == '':
        return True
    
    # For absolute URLs, require exact host match including port
    # and only allow HTTP/HTTPS schemes
    if test_url.scheme not in ('http', 'https'):
        return False
    
    return test_url.netloc == ref_url.netloc

main = Blueprint('main', __name__)


@main.route('/')
def index():
    """Home page"""
    return render_template('index.html')


@main.route('/register', methods=['GET', 'POST'])
def register():
    """User registration with role selection and key generation"""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        user = None
        keystore = None
        try:
            # Import required modules
            from app.asymmetric_crypto import AsymmetricCrypto
            from app.key_store import KeyStore
            from app.models import UserRole, CryptoLog
            from app.error_handlers import handle_mongodb_error, handle_crypto_error
            from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError, PyMongoError
            from datetime import datetime
            
            # Create user with role
            role_value = UserRole.ORGANIZATION if form.role.data == 'organization' else UserRole.CONSULTANT
            user = User(
                username=form.username.data,
                email=form.email.data,
                role=role_value
            )
            user.set_password(form.password.data)
            
            # Generate RSA key pair
            try:
                public_key, private_key = AsymmetricCrypto.generate_rsa_keypair()
            except ValueError as e:
                user_msg, log_msg = handle_crypto_error(e, "key generation")
                flash('Registration failed: ' + user_msg, 'danger')
                return render_template('register.html', form=form)
            except Exception as e:
                flash('Registration failed: Failed to generate encryption keys. Please try again.', 'danger')
                current_app.logger.error(f'Key generation error: {str(e)}')
                return render_template('register.html', form=form)
            
            # Store public key in SQLite User table
            user.public_key = public_key
            user.public_key_fingerprint = AsymmetricCrypto.get_public_key_fingerprint(public_key)
            user.key_generated_at = datetime.utcnow()
            
            # Add user to database first to get user ID
            db.session.add(user)
            db.session.flush()  # Get user.id without committing
            
            # Encrypt and store private key in MongoDB
            encrypted_key, salt, nonce = AsymmetricCrypto.encrypt_private_key(
                private_key, form.password.data
            )
            
            try:
                keystore = KeyStore()
                keystore.store_private_key(
                    user_id=user.id,
                    encrypted_key=encrypted_key,
                    salt=salt,
                    nonce=nonce,
                    metadata={'algorithm': 'RSA-2048', 'key_size': 2048}
                )
            except (ConnectionFailure, ServerSelectionTimeoutError, PyMongoError) as e:
                user_msg, log_msg = handle_mongodb_error(e, "private key storage")
                flash('Registration failed: ' + user_msg, 'danger')
                db.session.rollback()
                return render_template('register.html', form=form)
            finally:
                if keystore:
                    keystore.close()
            
            # Log key generation operation
            log_crypto_operation(
                user_id=user.id,
                operation='keypair_generated',
                details=f'RSA-2048 key pair generated for user {user.username}',
                success=True,
                ip_address=request.remote_addr
            )
            
            # Commit all changes
            db.session.commit()
            
            flash('Registration successful! Your encryption keys have been generated. Please log in.', 'success')
            return redirect(url_for('main.login'))
            
        except Exception as e:
            # Rollback on any error
            db.session.rollback()
            
            # Log the failure if we have a user ID
            if user and hasattr(user, 'id') and user.id:
                try:
                    log_crypto_operation(
                        user_id=user.id,
                        operation='keypair_generated',
                        details=f'Failed to generate key pair for user {user.username}',
                        success=False,
                        error_message=str(e),
                        ip_address=request.remote_addr
                    )
                except:
                    pass
            
            # Clean up MongoDB if user was created
            if user and hasattr(user, 'id') and user.id:
                try:
                    cleanup_keystore = KeyStore()
                    cleanup_keystore.delete_private_key(user.id)
                    cleanup_keystore.close()
                except:
                    pass
            
            flash('Registration failed. Please try again. If the problem persists, contact support.', 'danger')
            current_app.logger.error(f'Registration error: {str(e)}')
    
    return render_template('register.html', form=form)


@main.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    form = LoginForm()
    if request.method == 'POST':
        if form.validate_on_submit():
            user = User.query.filter_by(username=form.username.data).first()
            if user and user.check_password(form.password.data):
                login_user(user)
                flash('Login successful!', 'success')
                next_page = request.args.get('next')
                # Validate next URL to prevent open redirect vulnerabilities
                if next_page and is_safe_url(next_page):
                    return redirect(next_page)
                return redirect(url_for('main.dashboard'))
            else:
                flash('Invalid username or password.', 'danger')
        else:
            # Log form errors for debugging
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f'{field}: {error}', 'danger')
    
    return render_template('login.html', form=form)


@main.route('/logout')
@login_required
def logout():
    """User logout"""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('main.index'))


@main.route('/dashboard')
@login_required
def dashboard():
    """User dashboard showing uploaded and shared files"""
    # Get user's files
    user_files = EncryptedFile.query.filter_by(user_id=current_user.id).order_by(EncryptedFile.uploaded_at.desc()).all()
    
    # Get files shared with user
    shared_file_ids = [sf.file_id for sf in SharedFile.query.filter_by(recipient_id=current_user.id).all()]
    shared_files = EncryptedFile.query.filter(EncryptedFile.id.in_(shared_file_ids)).all() if shared_file_ids else []
    
    # Role-specific data
    role_data = {}
    approved_files_list = []
    
    if current_user.role == UserRole.ORGANIZATION:
        # Get pending access requests count for organization
        org_file_ids = [f.id for f in user_files]
        pending_requests = AccessRequest.query.filter(
            AccessRequest.file_id.in_(org_file_ids) if org_file_ids else False,
            AccessRequest.status == 'pending'
        ).count()
        
        approved_requests = AccessRequest.query.filter(
            AccessRequest.file_id.in_(org_file_ids) if org_file_ids else False,
            AccessRequest.status == 'approved'
        ).count()
        
        role_data = {
            'pending_requests': pending_requests,
            'approved_requests': approved_requests,
            'total_requests': pending_requests + approved_requests
        }
    
    elif current_user.role == UserRole.CONSULTANT:
        # Get consultant's access requests
        my_requests = AccessRequest.query.filter_by(consultant_id=current_user.id).all()
        
        pending_count = sum(1 for r in my_requests if r.status == 'pending')
        approved_count = sum(1 for r in my_requests if r.status == 'approved')
        denied_count = sum(1 for r in my_requests if r.status == 'denied')
        
        role_data = {
            'pending_requests': pending_count,
            'approved_requests': approved_count,
            'denied_requests': denied_count,
            'total_requests': len(my_requests)
        }
        
        # Get approved files for consultant dashboard
        approved_requests_list = AccessRequest.query.filter_by(
            consultant_id=current_user.id,
            status='approved'
        ).order_by(AccessRequest.processed_at.desc()).limit(5).all()
        
        approved_files_list = [req.file for req in approved_requests_list if req.file]
    
    return render_template('dashboard.html', 
                         user_files=user_files, 
                         shared_files=shared_files,
                         role_data=role_data,
                         approved_files_list=approved_files_list)


@main.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    """File upload and encryption"""
    form = FileUploadForm()
    
    if form.validate_on_submit():
        file = form.file.data
        algorithm = form.algorithm.data
        
        if file and allowed_file(file.filename):
            # Secure the filename
            original_filename = secure_filename(file.filename)
            file_id = generate_file_id()
            
            # Save file temporarily
            temp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'{file_id}_{original_filename}')
            file.save(temp_path)
            
            try:
                # Encrypt the file — CryptoEngine now returns a wrapped DEK blob instead of raw key
                encrypted_data, wrapped_key, iv, exec_time, original_size, encrypted_size = CryptoEngine.encrypt_file(
                    temp_path, algorithm
                )
                
                # Save encrypted file
                encrypted_filename = f'{file_id}_encrypted'
                encrypted_path = os.path.join(current_app.config['ENCRYPTED_FOLDER'], encrypted_filename)
                
                with open(encrypted_path, 'wb') as f:
                    f.write(encrypted_data)
                
                # Store metadata in database
                file_type = get_file_type(original_filename)
                encrypted_file = EncryptedFile(
                    file_id=file_id,
                    filename=encrypted_filename,
                    original_filename=original_filename,
                    file_type=file_type,
                    file_size=original_size,
                    encrypted_path=encrypted_path,
                    algorithm=algorithm,
                    # store the wrapped DEK blob in new column
                    wrapped_key=wrapped_key,
                    wrapped_key_version='V1',
                    # keep legacy column None to avoid confusion
                    encryption_key=None,
                    iv=iv,
                    user_id=current_user.id
                )
                db.session.add(encrypted_file)
                db.session.commit()
                
                # Store performance metrics
                metric = PerformanceMetric(
                    file_id=encrypted_file.id,
                    algorithm=algorithm,
                    operation='encryption',
                    data_type=file_type,
                    execution_time=exec_time,
                    input_size=original_size,
                    output_size=encrypted_size
                )
                db.session.add(metric)
                db.session.commit()
                
                # Log operation
                log_operation(current_user.id, file_id, 'upload', algorithm, True, None, request.remote_addr)
                
                # Parse financial report if it's an Excel file
                if file_type == 'spreadsheet':
                    financial_data = parse_financial_report(temp_path)
                    if financial_data:
                        # Encrypt financial data (uses same wrapped_key produced earlier)
                        encrypted_dict, _, _ = encrypt_financial_data(financial_data, algorithm, wrapped_key)
                        
                        # Store in database
                        report = FinancialReport(
                            file_id=encrypted_file.id,
                            **encrypted_dict
                        )
                        db.session.add(report)
                        db.session.commit()
                
                # Remove temporary file
                os.remove(temp_path)
                
                flash(f'File encrypted successfully using {algorithm}!', 'success')
                return redirect(url_for('main.file_detail', file_id=file_id))
                
            except Exception as e:
                db.session.rollback()
                log_operation(current_user.id, file_id, 'upload', algorithm, False, str(e), request.remote_addr)
                flash(f'Error encrypting file: {str(e)}', 'danger')
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        else:
            flash('Invalid file type.', 'danger')
    
    return render_template('upload.html', form=form)


@main.route('/financial-report', methods=['GET', 'POST'])
@login_required
def financial_report():
    """Create financial report from form"""
    form = FinancialReportForm()
    
    if form.validate_on_submit():
        algorithm = form.algorithm.data
        
        # Collect form data
        financial_data = {
            'company_name': form.company_name.data,
            'report_period': form.report_period.data,
            'department': form.department.data,
            'total_revenue': str(form.total_revenue.data),
            'total_expenses': str(form.total_expenses.data),
            'net_profit': str(form.net_profit.data) if form.net_profit.data else '0',
            'budget_allocated': str(form.budget_allocated.data),
            'budget_spent': str(form.budget_spent.data),
            'variance': str(form.variance.data) if form.variance.data else '0',
            'notes': form.notes.data if form.notes.data else ''
        }
        
        try:
            # Generate file ID
            file_id = generate_file_id()
            
            # Create Excel file
            temp_filename = f'{file_id}_report.xlsx'
            temp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], temp_filename)
            create_financial_report_template(temp_path)
            
            # Populate the template with data
            import openpyxl
            wb = openpyxl.load_workbook(temp_path)
            ws = wb.active
            
            ws['B3'] = financial_data['company_name']
            ws['B4'] = financial_data['report_period']
            ws['B5'] = financial_data['department']
            ws['B8'] = float(financial_data['total_revenue'])
            ws['B9'] = float(financial_data['total_expenses'])
            ws['B13'] = float(financial_data['budget_allocated'])
            ws['B14'] = float(financial_data['budget_spent'])
            ws['A18'] = financial_data['notes']
            
            wb.save(temp_path)
            
            # Encrypt the file — returns encrypted bytes + wrapped DEK
            encrypted_data, wrapped_key, iv, exec_time, original_size, encrypted_size = CryptoEngine.encrypt_file(
                temp_path, algorithm
            )
            
            # Save encrypted file
            encrypted_filename = f'{file_id}_encrypted'
            encrypted_path = os.path.join(current_app.config['ENCRYPTED_FOLDER'], encrypted_filename)
            
            with open(encrypted_path, 'wb') as f:
                f.write(encrypted_data)
            
            # Store metadata
            encrypted_file = EncryptedFile(
                file_id=file_id,
                filename=encrypted_filename,
                original_filename=temp_filename,
                file_type='spreadsheet',
                file_size=original_size,
                encrypted_path=encrypted_path,
                algorithm=algorithm,
                    # store wrapped DEK in new column
                    wrapped_key=wrapped_key,
                    wrapped_key_version='V1',
                    encryption_key=None,
                iv=iv,
                user_id=current_user.id
            )
            db.session.add(encrypted_file)
            db.session.commit()
            
            # Store performance metrics
            metric = PerformanceMetric(
                file_id=encrypted_file.id,
                algorithm=algorithm,
                operation='encryption',
                data_type='numerical',
                execution_time=exec_time,
                input_size=original_size,
                output_size=encrypted_size
            )
            db.session.add(metric)
            
            # Encrypt and store financial data (use wrapped_key)
            encrypted_dict, _, _ = encrypt_financial_data(financial_data, algorithm, wrapped_key)
            report = FinancialReport(
                file_id=encrypted_file.id,
                **encrypted_dict
            )
            db.session.add(report)
            db.session.commit()
            
            # Log operation
            log_operation(current_user.id, file_id, 'upload', algorithm, True, None, request.remote_addr)
            
            # Remove temporary file
            os.remove(temp_path)
            
            flash(f'Financial report created and encrypted successfully using {algorithm}!', 'success')
            return redirect(url_for('main.file_detail', file_id=file_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating financial report: {str(e)}', 'danger')
    
    return render_template('financial_report_form.html', form=form)


@main.route('/file/<file_id>')
@login_required
def file_detail(file_id):
    """View file details"""
    file = EncryptedFile.query.filter_by(file_id=file_id).first_or_404()
    
    # Check if user has permission to view
    approved_request = None
    if file.user_id != current_user.id:
        # Check for shared file access (legacy)
        shared = SharedFile.query.filter_by(file_id=file.id, recipient_id=current_user.id).first()
        
        # Check for approved access request (consultant access)
        if current_user.role == UserRole.CONSULTANT:
            approved_request = AccessRequest.query.filter_by(
                consultant_id=current_user.id,
                file_id=file.id,
                status='approved'
            ).first()
        
        # Deny access if neither shared nor approved request exists
        if not shared and not approved_request:
            flash('You do not have permission to view this file.', 'danger')
            return redirect(url_for('main.dashboard'))
    
    # Get performance metrics
    metrics = PerformanceMetric.query.filter_by(file_id=file.id).all()
    
    # Get financial report if exists
    financial_report = FinancialReport.query.filter_by(file_id=file.id).first()
    decrypted_report = None
    
    if financial_report:
        try:
            encrypted_dict = {
                'encrypted_company_name': financial_report.encrypted_company_name,
                'encrypted_report_period': financial_report.encrypted_report_period,
                'encrypted_department': financial_report.encrypted_department,
                'encrypted_total_revenue': financial_report.encrypted_total_revenue,
                'encrypted_total_expenses': financial_report.encrypted_total_expenses,
                'encrypted_net_profit': financial_report.encrypted_net_profit,
                'encrypted_budget_allocated': financial_report.encrypted_budget_allocated,
                'encrypted_budget_spent': financial_report.encrypted_budget_spent,
                'encrypted_variance': financial_report.encrypted_variance,
                'encrypted_notes': financial_report.encrypted_notes,
            }
            # Prefer wrapped_key column, fall back to legacy encryption_key
            wrapped = file.wrapped_key if getattr(file, 'wrapped_key', None) else file.encryption_key
            decrypted_report = decrypt_financial_data(encrypted_dict, file.algorithm, wrapped, file.iv)
        except Exception as e:
            flash(f'Error decrypting financial data: {str(e)}', 'warning')
    
    # Get shared users
    shared_with = db.session.query(User).join(SharedFile).filter(SharedFile.file_id == file.id).all()
    
    return render_template('file_detail.html', 
                         file=file, 
                         metrics=metrics, 
                         financial_report=decrypted_report,
                         shared_with=shared_with,
                         approved_request=approved_request,
                         format_size=format_file_size,
                         format_time=format_execution_time)


@main.route('/file/<file_id>/download')
@login_required
def download_file(file_id):
    """Download and decrypt file"""
    from flask import session
    from app.asymmetric_crypto import AsymmetricCrypto
    from app.crypto_utils import CryptoEngine
    from Crypto.Cipher import AES
    import time
    
    file = EncryptedFile.query.filter_by(file_id=file_id).first_or_404()
    
    # Check permission
    if file.user_id != current_user.id:
        # Check for shared file access (legacy)
        shared = SharedFile.query.filter_by(file_id=file.id, recipient_id=current_user.id).first()
        
        # Check for approved access request (consultant access)
        approved_request = None
        if current_user.role == UserRole.CONSULTANT:
            approved_request = AccessRequest.query.filter_by(
                consultant_id=current_user.id,
                file_id=file.id,
                status='approved'
            ).first()
        
        # Deny access if neither shared nor approved request exists
        if not shared and not approved_request:
            flash('You do not have permission to download this file.', 'danger')
            return redirect(url_for('main.dashboard'))
        
        # For shared files, check download permission
        if shared and not shared.can_download:
            flash('You do not have permission to download this file.', 'danger')
            return redirect(url_for('main.dashboard'))
        
        # For consultant access, check if request is revoked
        if approved_request and approved_request.status == 'revoked':
            flash('Your access to this file has been revoked.', 'danger')
            return redirect(url_for('main.my_requests'))
    
    try:
        # Read encrypted file
        with open(file.encrypted_path, 'rb') as f:
            encrypted_data = f.read()
        
        # Determine which wrapped key to use
        wrapped_key = None
        
        # If user is consultant with approved access, use the RSA-wrapped key from session
        if file.user_id != current_user.id and current_user.role == UserRole.CONSULTANT:
            approved_request = AccessRequest.query.filter_by(
                consultant_id=current_user.id,
                file_id=file.id,
                status='approved'
            ).first()
            
            if approved_request:
                # Check if symmetric key is in session
                session_key = f'symmetric_key_{approved_request.id}'
                
                if session_key not in session:
                    flash('Please decrypt the symmetric key first before downloading.', 'warning')
                    return redirect(url_for('main.decrypt_key', request_id=approved_request.id))
                
                # Retrieve symmetric key from session (stored as hex string)
                symmetric_key_hex = session[session_key]
                symmetric_key = bytes.fromhex(symmetric_key_hex)
                
                # For consultant access, we need to wrap the symmetric key in the format
                # expected by CryptoEngine.decrypt_file (which expects a KEK-wrapped key)
                # Since we have the raw symmetric key, we need to use it directly
                # We'll need to modify our approach here
                
                # Actually, looking at the decrypt_file signature, it expects wrapped_dek
                # which gets unwrapped by the KEK. For consultant access, we have the
                # raw DEK, so we need to handle this differently.
                
                # Decrypt directly using the symmetric key
                from Crypto.Cipher import DES, ARC4
                from Crypto.Util.Padding import unpad
                
                start_time = time.time()
                
                if file.algorithm.upper() == 'AES':
                    cipher = AES.new(symmetric_key, AES.MODE_CBC, file.iv)
                    decrypted_padded = cipher.decrypt(encrypted_data)
                    decrypted_data = unpad(decrypted_padded, 16)  # AES block size
                    
                elif file.algorithm.upper() == 'DES':
                    cipher = DES.new(symmetric_key, DES.MODE_CBC, file.iv)
                    decrypted_padded = cipher.decrypt(encrypted_data)
                    decrypted_data = unpad(decrypted_padded, 8)  # DES block size
                    
                elif file.algorithm.upper() == 'RC4':
                    cipher = ARC4.new(symmetric_key)
                    decrypted_data = cipher.decrypt(encrypted_data)
                    
                else:
                    flash(f'Algorithm {file.algorithm} not yet supported for consultant access.', 'danger')
                    return redirect(url_for('main.my_requests'))
                
                exec_time = time.time() - start_time
        else:
            # Owner or shared file access - use KEK-wrapped key
            wrapped_key = file.wrapped_key if getattr(file, 'wrapped_key', None) else file.encryption_key
            decrypted_data, exec_time = CryptoEngine.decrypt_file(
                encrypted_data, file.algorithm, wrapped_key, file.iv
            )
        
        # Store decryption performance metric
        metric = PerformanceMetric(
            file_id=file.id,
            algorithm=file.algorithm,
            operation='decryption',
            data_type=file.file_type,
            execution_time=exec_time,
            input_size=len(encrypted_data),
            output_size=len(decrypted_data)
        )
        db.session.add(metric)
        db.session.commit()
        
        # Log operation
        log_operation(current_user.id, file_id, 'download', file.algorithm, True, None, request.remote_addr)
        
        # Send file
        return send_file(
            BytesIO(decrypted_data),
            as_attachment=True,
            download_name=file.original_filename,
            mimetype='application/octet-stream'
        )
        
    except Exception as e:
        log_operation(current_user.id, file_id, 'download', file.algorithm, False, str(e), request.remote_addr)
        flash(f'Error downloading file: {str(e)}', 'danger')
        return redirect(url_for('main.file_detail', file_id=file_id))


@main.route('/file/<file_id>/share', methods=['GET', 'POST'])
@login_required
def share_file(file_id):
    """Share file with another user"""
    file = EncryptedFile.query.filter_by(file_id=file_id).first_or_404()
    
    # Only owner can share
    if file.user_id != current_user.id:
        flash('You do not have permission to share this file.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    form = ShareFileForm()
    
    if form.validate_on_submit():
        recipient = User.query.filter_by(username=form.recipient_username.data).first()
        
        if recipient.id == current_user.id:
            flash('You cannot share a file with yourself.', 'warning')
        else:
            # Check if already shared
            existing_share = SharedFile.query.filter_by(file_id=file.id, recipient_id=recipient.id).first()
            
            if existing_share:
                flash('File is already shared with this user.', 'info')
            else:
                # Create share
                share = SharedFile(file_id=file.id, recipient_id=recipient.id)
                db.session.add(share)
                db.session.commit()
                
                # Log operation
                log_operation(current_user.id, file_id, 'share', file.algorithm, True, None, request.remote_addr)
                
                flash(f'File shared successfully with {recipient.username}!', 'success')
                return redirect(url_for('main.file_detail', file_id=file_id))
    
    return render_template('share_file.html', form=form, file=file)


@main.route('/file/<file_id>/delete', methods=['POST'])
@login_required
def delete_file(file_id):
    """Delete file"""
    file = EncryptedFile.query.filter_by(file_id=file_id).first_or_404()
    
    # Only owner can delete
    if file.user_id != current_user.id:
        flash('You do not have permission to delete this file.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Delete encrypted file from storage
        if os.path.exists(file.encrypted_path):
            os.remove(file.encrypted_path)
        
        # Delete from database (cascades to related records)
        db.session.delete(file)
        db.session.commit()
        
        flash('File deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting file: {str(e)}', 'danger')
    
    return redirect(url_for('main.dashboard'))


@main.route('/performance')
@login_required
def performance():
    """View performance comparison"""
    # Get all metrics for user's files
    user_file_ids = [f.id for f in EncryptedFile.query.filter_by(user_id=current_user.id).all()]
    metrics = PerformanceMetric.query.filter(PerformanceMetric.file_id.in_(user_file_ids)).all() if user_file_ids else []
    
    # Organize metrics by algorithm and operation
    performance_data = {
        'AES': {'encryption': [], 'decryption': []},
        'DES': {'encryption': [], 'decryption': []},
        'RC4': {'encryption': [], 'decryption': []}
    }
    
    for metric in metrics:
        if metric.algorithm in performance_data and metric.operation in performance_data[metric.algorithm]:
            performance_data[metric.algorithm][metric.operation].append(metric)
    
    return render_template('performance.html', 
                         metrics=metrics,
                         performance_data=performance_data,
                         format_size=format_file_size,
                         format_time=format_execution_time)


@main.route('/download-template')
@login_required
def download_template():
    """Download financial report template"""
    try:
        # Create template
        temp_filename = 'Financial_Report_Template.xlsx'
        temp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], temp_filename)
        create_financial_report_template(temp_path)
        
        # Send file
        return send_file(
            temp_path,
            as_attachment=True,
            download_name=temp_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error creating template: {str(e)}', 'danger')
        return redirect(url_for('main.dashboard'))


@main.route('/shared-files')
@login_required
def shared_files():
    """View files shared with user"""
    shared_file_ids = [sf.file_id for sf in SharedFile.query.filter_by(recipient_id=current_user.id).all()]
    files = EncryptedFile.query.filter(EncryptedFile.id.in_(shared_file_ids)).all() if shared_file_ids else []
    
    return render_template('shared_files.html', files=files, format_size=format_file_size)


@main.route('/organizations')
@login_required
def view_organizations():
    """Consultant views organizations and their files"""
    from app.decorators import consultant_required
    
    # Check if user is consultant
    if current_user.role != UserRole.CONSULTANT:
        flash('Access denied. This page is only available to consultant users.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get all organization users
    organizations = User.query.filter_by(role=UserRole.ORGANIZATION).all()
    
    # Get files for each organization with request status
    org_data = []
    for org in organizations:
        files = EncryptedFile.query.filter_by(user_id=org.id).all()
        
        # Check existing requests for each file
        files_with_status = []
        for file in files:
            existing_request = AccessRequest.query.filter_by(
                consultant_id=current_user.id,
                file_id=file.id
            ).first()
            
            files_with_status.append({
                'file': file,
                'request': existing_request
            })
        
        org_data.append({
            'organization': org,
            'files': files_with_status
        })
    
    return render_template('organizations.html', org_data=org_data, format_size=format_file_size)


@main.route('/request-access/<int:file_id>', methods=['GET', 'POST'])
@login_required
def request_access(file_id):
    """Consultant submits access request"""
    from app.decorators import consultant_required
    
    # Check if user is consultant
    if current_user.role != UserRole.CONSULTANT:
        flash('Access denied. Only consultants can request access to files.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get the file
    file = EncryptedFile.query.get_or_404(file_id)
    
    # Check for existing request (duplicate prevention)
    existing_request = AccessRequest.query.filter_by(
        consultant_id=current_user.id,
        file_id=file.id
    ).first()
    
    if existing_request:
        flash(f'You already have a {existing_request.status} request for this file.', 'info')
        return redirect(url_for('main.view_organizations'))
    
    form = AccessRequestForm()
    
    if form.validate_on_submit():
        try:
            # Create access request
            access_request = AccessRequest(
                consultant_id=current_user.id,
                organization_id=file.user_id,
                file_id=file.id,
                status='pending'
            )
            
            db.session.add(access_request)
            db.session.commit()
            
            flash(f'Access request submitted successfully for file: {file.original_filename}', 'success')
            return redirect(url_for('main.view_organizations'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error submitting access request: {str(e)}', 'danger')
            current_app.logger.error(f'Access request error: {str(e)}')
    
    return render_template('request_access.html', form=form, file=file)


@main.route('/my-requests')
@login_required
def my_requests():
    """Consultant views their access requests"""
    from app.decorators import consultant_required
    
    # Check if user is consultant
    if current_user.role != UserRole.CONSULTANT:
        flash('Access denied. This page is only available to consultant users.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get all requests submitted by this consultant
    requests = AccessRequest.query.filter_by(
        consultant_id=current_user.id
    ).order_by(AccessRequest.requested_at.desc()).all()
    
    return render_template('my_requests.html', requests=requests, format_size=format_file_size)


@main.route('/access-requests')
@login_required
def view_access_requests():
    """Organization views pending access requests for their files"""
    from app.decorators import organization_required
    
    # Check if user is organization
    if current_user.role != UserRole.ORGANIZATION:
        flash('Access denied. This page is only available to organization users.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get all files owned by this organization
    org_files = EncryptedFile.query.filter_by(user_id=current_user.id).all()
    org_file_ids = [f.id for f in org_files]
    
    # Get status filter from query parameters (optional)
    status_filter = request.args.get('status', 'pending')
    
    # Get all access requests for this organization's files
    if status_filter and status_filter != 'all':
        requests = AccessRequest.query.filter(
            AccessRequest.file_id.in_(org_file_ids) if org_file_ids else False,
            AccessRequest.status == status_filter
        ).order_by(AccessRequest.requested_at.desc()).all()
    else:
        requests = AccessRequest.query.filter(
            AccessRequest.file_id.in_(org_file_ids) if org_file_ids else False
        ).order_by(AccessRequest.requested_at.desc()).all()
    
    return render_template('access_requests.html', 
                         requests=requests, 
                         status_filter=status_filter,
                         format_size=format_file_size)


@main.route('/approve-request/<int:request_id>', methods=['POST'])
@login_required
def approve_request(request_id):
    """Organization approves access request and wraps key"""
    from app.decorators import organization_required
    from app.asymmetric_crypto import AsymmetricCrypto
    from app.crypto_utils import _unwrap_key
    from datetime import datetime
    
    # Check if user is organization
    if current_user.role != UserRole.ORGANIZATION:
        flash('Access denied. Only organizations can approve access requests.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get the access request
    access_request = AccessRequest.query.get_or_404(request_id)
    
    # Verify the request is for a file owned by this organization
    if access_request.organization_id != current_user.id:
        flash('Access denied. You can only approve requests for your own files.', 'danger')
        return redirect(url_for('main.view_access_requests'))
    
    # Verify request is in pending status
    if access_request.status != 'pending':
        flash(f'Cannot approve request with status: {access_request.status}', 'warning')
        return redirect(url_for('main.view_access_requests'))
    
    try:
        from app.error_handlers import handle_crypto_error
        
        # Get the file
        file = EncryptedFile.query.get(access_request.file_id)
        if not file:
            flash('File not found.', 'danger')
            return redirect(url_for('main.view_access_requests'))
        
        # Retrieve file's wrapped symmetric key from database
        wrapped_dek = file.wrapped_key
        if not wrapped_dek:
            # Fallback to legacy encryption_key if wrapped_key is not available
            wrapped_dek = file.encryption_key
            if not wrapped_dek:
                flash('No encryption key found for this file.', 'danger')
                
                # Log the error
                try:
                    log_crypto_operation(
                        user_id=current_user.id,
                        operation='key_wrapped',
                        details=f'No encryption key found for file {file.original_filename}',
                        success=False,
                        error_message='File has no encryption key',
                        ip_address=request.remote_addr
                    )
                except:
                    pass
                
                return redirect(url_for('main.view_access_requests'))
        
        # Unwrap symmetric key using KEK
        # The wrapped_dek is already KEK-wrapped, so we unwrap it to get the plaintext DEK
        try:
            symmetric_key = _unwrap_key(wrapped_dek)
        except Exception as e:
            flash('Failed to unwrap file encryption key. The file may be corrupted.', 'danger')
            current_app.logger.error(f'Key unwrapping error: {str(e)}')
            
            # Log the error
            try:
                log_crypto_operation(
                    user_id=current_user.id,
                    operation='key_wrapped',
                    details=f'Failed to unwrap file encryption key for approval',
                    success=False,
                    error_message=str(e),
                    ip_address=request.remote_addr
                )
            except:
                pass
            
            return redirect(url_for('main.view_access_requests'))
        
        # Retrieve consultant's public key
        consultant = User.query.get(access_request.consultant_id)
        if not consultant:
            flash('Consultant not found.', 'danger')
            return redirect(url_for('main.view_access_requests'))
        
        if not consultant.public_key:
            flash(f'Consultant {consultant.username} does not have a public key. They may need to re-register.', 'danger')
            return redirect(url_for('main.view_access_requests'))
        
        # Wrap symmetric key with consultant's RSA public key
        try:
            wrapped_symmetric_key = AsymmetricCrypto.wrap_symmetric_key(
                symmetric_key, consultant.public_key
            )
        except ValueError as e:
            user_msg, log_msg = handle_crypto_error(e, "key wrapping")
            flash(user_msg, 'danger')
            
            # Log the error
            try:
                log_crypto_operation(
                    user_id=current_user.id,
                    operation='key_wrapped',
                    details=f'Failed to wrap key for consultant {consultant.username}',
                    success=False,
                    error_message=str(e),
                    ip_address=request.remote_addr
                )
            except:
                pass
            
            return redirect(url_for('main.view_access_requests'))
        
        # Store wrapped key in AccessRequest record
        access_request.wrapped_symmetric_key = wrapped_symmetric_key
        
        # Update request status to approved
        access_request.status = 'approved'
        
        # Record processed_at timestamp
        access_request.processed_at = datetime.utcnow()
        
        # Log key wrapping operation
        log_crypto_operation(
            user_id=current_user.id,
            operation='key_wrapped',
            details=f'Wrapped symmetric key for consultant {consultant.username} to access file {file.original_filename}',
            success=True,
            ip_address=request.remote_addr
        )
        
        # Commit all changes
        db.session.commit()
        
        flash(f'Access request approved for {consultant.username}', 'success')
        
    except Exception as e:
        # Rollback on failure
        db.session.rollback()
        
        # Log the error
        try:
            log_crypto_operation(
                user_id=current_user.id,
                operation='key_wrapped',
                details=f'Failed to approve access request {request_id}',
                success=False,
                error_message=str(e),
                ip_address=request.remote_addr
            )
        except:
            pass
        
        flash('An unexpected error occurred while approving the request. Please try again.', 'danger')
        current_app.logger.error(f'Access request approval error: {str(e)}')
    
    return redirect(url_for('main.view_access_requests'))



@main.route('/deny-request/<int:request_id>', methods=['POST'])
@login_required
def deny_request(request_id):
    """Organization denies access request"""
    from app.decorators import organization_required
    from datetime import datetime
    
    # Check if user is organization
    if current_user.role != UserRole.ORGANIZATION:
        flash('Access denied. Only organizations can deny access requests.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get the access request
    access_request = AccessRequest.query.get_or_404(request_id)
    
    # Verify the request is for a file owned by this organization
    if access_request.organization_id != current_user.id:
        flash('Access denied. You can only deny requests for your own files.', 'danger')
        return redirect(url_for('main.view_access_requests'))
    
    # Verify request is in pending status
    if access_request.status != 'pending':
        flash(f'Cannot deny request with status: {access_request.status}', 'warning')
        return redirect(url_for('main.view_access_requests'))
    
    try:
        # Get the file and consultant for logging
        file = EncryptedFile.query.get(access_request.file_id)
        consultant = User.query.get(access_request.consultant_id)
        
        # Update request status to denied
        access_request.status = 'denied'
        
        # Record processed_at timestamp
        access_request.processed_at = datetime.utcnow()
        
        # Log denial operation
        log_crypto_operation(
            user_id=current_user.id,
            operation='access_denied',
            details=f'Denied access request from consultant {consultant.username if consultant else "unknown"} for file {file.original_filename if file else "unknown"}',
            success=True,
            ip_address=request.remote_addr
        )
        
        # Commit all changes
        db.session.commit()
        
        flash(f'Access request denied for {consultant.username if consultant else "consultant"}', 'success')
        
    except Exception as e:
        # Rollback on failure
        db.session.rollback()
        
        # Log the error
        try:
            log_crypto_operation(
                user_id=current_user.id,
                operation='access_denied',
                details=f'Failed to deny access request {request_id}',
                success=False,
                error_message=str(e),
                ip_address=request.remote_addr
            )
        except:
            pass
        
        flash(f'Error denying access request: {str(e)}', 'danger')
        current_app.logger.error(f'Access request denial error: {str(e)}')
    
    return redirect(url_for('main.view_access_requests'))


@main.route('/revoke-access/<int:request_id>', methods=['POST'])
@login_required
def revoke_access(request_id):
    """Organization revokes access to a file"""
    from app.decorators import organization_required
    from datetime import datetime
    
    # Check if user is organization
    if current_user.role != UserRole.ORGANIZATION:
        flash('Access denied. Only organizations can revoke access.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get the access request
    access_request = AccessRequest.query.get_or_404(request_id)
    
    # Verify the request is for a file owned by this organization
    if access_request.organization_id != current_user.id:
        flash('Access denied. You can only revoke access for your own files.', 'danger')
        return redirect(url_for('main.view_access_requests'))
    
    # Verify request is not already revoked or denied
    if access_request.status in ['revoked', 'denied']:
        flash(f'Cannot revoke request with status: {access_request.status}.', 'warning')
        return redirect(url_for('main.view_access_requests'))
    
    try:
        # Get the file and consultant for logging
        file = EncryptedFile.query.get(access_request.file_id)
        consultant = User.query.get(access_request.consultant_id)
        
        # Update request status to revoked
        access_request.status = 'revoked'
        
        # Delete wrapped symmetric key from request
        access_request.wrapped_symmetric_key = None
        
        # Record processed_at timestamp
        access_request.processed_at = datetime.utcnow()
        
        # Invalidate any cached keys in consultant sessions
        # Note: Session invalidation would require additional session management
        # For now, we rely on the status check in download route
        
        # Log revocation operation
        log_crypto_operation(
            user_id=current_user.id,
            operation='access_revoked',
            details=f'Revoked access for consultant {consultant.username if consultant else "unknown"} to file {file.original_filename if file else "unknown"}',
            success=True,
            ip_address=request.remote_addr
        )
        
        # Commit all changes
        db.session.commit()
        
        flash(f'Access revoked for {consultant.username if consultant else "consultant"}', 'success')
        
    except Exception as e:
        # Rollback on failure
        db.session.rollback()
        
        # Log the error
        try:
            log_crypto_operation(
                user_id=current_user.id,
                operation='access_revoked',
                details=f'Failed to revoke access request {request_id}',
                success=False,
                error_message=str(e),
                ip_address=request.remote_addr
            )
        except:
            pass
        
        flash(f'Error revoking access: {str(e)}', 'danger')
        current_app.logger.error(f'Access revocation error: {str(e)}')
    
    return redirect(url_for('main.view_access_requests'))


@main.route('/decrypt-key/<int:request_id>', methods=['GET', 'POST'])
@login_required
def decrypt_key(request_id):
    """Consultant decrypts symmetric key using their private key"""
    from app.decorators import consultant_required
    from app.forms import DecryptKeyForm
    from app.asymmetric_crypto import AsymmetricCrypto
    from app.key_store import KeyStore
    from datetime import datetime
    from flask import session
    
    # Check if user is consultant
    if current_user.role != UserRole.CONSULTANT:
        flash('Access denied. Only consultants can decrypt keys.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get the access request
    access_request = AccessRequest.query.get_or_404(request_id)
    
    # Verify the request belongs to this consultant
    if access_request.consultant_id != current_user.id:
        flash('Access denied. You can only decrypt keys for your own requests.', 'danger')
        return redirect(url_for('main.my_requests'))
    
    # Verify request is approved
    if access_request.status != 'approved':
        flash(f'Cannot decrypt key for request with status: {access_request.status}', 'warning')
        return redirect(url_for('main.my_requests'))
    
    # Verify wrapped symmetric key exists
    if not access_request.wrapped_symmetric_key:
        flash('No wrapped symmetric key found for this request.', 'danger')
        return redirect(url_for('main.my_requests'))
    
    form = DecryptKeyForm()
    
    if form.validate_on_submit():
        password = form.password.data
        keystore = None
        private_key_pem = None
        symmetric_key = None
        
        try:
            from app.error_handlers import handle_mongodb_error, handle_crypto_error
            from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError, PyMongoError
            
            # Retrieve encrypted private key from MongoDB
            try:
                keystore = KeyStore()
                key_data = keystore.retrieve_private_key(current_user.id)
                
                if not key_data:
                    raise ValueError('Private key not found in key store')
            except (ConnectionFailure, ServerSelectionTimeoutError, PyMongoError) as e:
                user_msg, log_msg = handle_mongodb_error(e, "private key retrieval")
                flash(user_msg, 'danger')
                return render_template('decrypt_key.html', form=form, access_request=access_request, file=access_request.file)
            
            # Decrypt private key using password
            try:
                private_key_pem = AsymmetricCrypto.decrypt_private_key(
                    encrypted_key=key_data['encrypted_key'],
                    password=password,
                    salt=key_data['salt'],
                    nonce=key_data['nonce']
                )
            except ValueError as e:
                user_msg, log_msg = handle_crypto_error(e, "private key decryption")
                
                # Log failed decryption - use 'key_unwrapped' as the overall operation being attempted
                try:
                    log_crypto_operation(
                        user_id=current_user.id,
                        operation='key_unwrapped',
                        details=f'Failed to decrypt private key for request {request_id}',
                        success=False,
                        error_message=str(e),
                        ip_address=request.remote_addr
                    )
                except:
                    pass
                
                flash(user_msg, 'danger')
                return render_template('decrypt_key.html', form=form, access_request=access_request, file=access_request.file)
            
            # Unwrap symmetric key using RSA private key
            try:
                symmetric_key = AsymmetricCrypto.unwrap_symmetric_key(
                    wrapped_key=access_request.wrapped_symmetric_key,
                    private_key_pem=private_key_pem
                )
            except ValueError as e:
                user_msg, log_msg = handle_crypto_error(e, "symmetric key unwrapping")
                
                # Log failed unwrapping
                try:
                    log_crypto_operation(
                        user_id=current_user.id,
                        operation='key_unwrapped',
                        details=f'Failed to unwrap symmetric key for request {request_id}',
                        success=False,
                        error_message=str(e),
                        ip_address=request.remote_addr
                    )
                except:
                    pass
                
                flash(user_msg, 'danger')
                return render_template('decrypt_key.html', form=form, access_request=access_request, file=access_request.file)
            
            # Store decrypted symmetric key in session (temporary)
            # Use a session key specific to this request
            session_key = f'symmetric_key_{request_id}'
            session[session_key] = symmetric_key.hex()  # Store as hex string
            
            # Log unwrapping operation
            log_crypto_operation(
                user_id=current_user.id,
                operation='key_unwrapped',
                details=f'Successfully unwrapped symmetric key for file {access_request.file.original_filename}',
                success=True,
                ip_address=request.remote_addr
            )
            
            flash('Symmetric key decrypted successfully! You can now download the file.', 'success')
            return redirect(url_for('main.my_requests'))
            
        except Exception as e:
            # Handle unexpected errors
            error_message = str(e)
            
            # Log failed decryption
            try:
                log_crypto_operation(
                    user_id=current_user.id,
                    operation='key_unwrapped',
                    details=f'Unexpected error unwrapping symmetric key for request {request_id}',
                    success=False,
                    error_message=error_message,
                    ip_address=request.remote_addr
                )
            except:
                pass
            
            flash('An unexpected error occurred. Please try again. If the problem persists, contact support.', 'danger')
            current_app.logger.error(f'Key decryption error: {error_message}')
            
        finally:
            # Clear private key from memory after use
            if private_key_pem is not None:
                private_key_pem = None
            if symmetric_key is not None:
                symmetric_key = None
            
            # Close keystore connection
            if keystore:
                keystore.close()
    
    # Get file information for display
    file = access_request.file
    
    return render_template('decrypt_key.html', form=form, access_request=access_request, file=file)


@main.route('/approved-files')
@login_required
def approved_files():
    """Consultant views files with approved access"""
    from app.decorators import consultant_required
    from flask import session
    
    # Check if user is consultant
    if current_user.role != UserRole.CONSULTANT:
        flash('Access denied. This page is only available to consultant users.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get all access requests for this consultant with approved or revoked status
    # We include revoked to show them but mark them appropriately
    approved_requests = AccessRequest.query.filter(
        AccessRequest.consultant_id == current_user.id,
        or_(AccessRequest.status == 'approved', AccessRequest.status == 'revoked')
    ).order_by(AccessRequest.processed_at.desc()).all()
    
    # Check decryption status for each request
    # A request is "decrypted" if the symmetric key is in the session
    decryption_status = {}
    for req in approved_requests:
        session_key = f'symmetric_key_{req.id}'
        decryption_status[req.id] = session_key in session
    
    return render_template('approved_files.html', 
                         approved_requests=approved_requests,
                         decryption_status=decryption_status,
                         format_size=format_file_size)



@main.route('/crypto-logs')
@login_required
def view_crypto_logs():
    """
    View cryptographic operation logs (admin interface).
    
    This route displays all cryptographic operations logged in the system,
    including key generation, wrapping, unwrapping, and access control events.
    
    For now, this is accessible to all authenticated users. In a production
    system, this should be restricted to administrators only.
    """
    from app.models import CryptoLog
    
    # Get filter parameters from query string
    operation_filter = request.args.get('operation', 'all')
    success_filter = request.args.get('success', 'all')
    user_filter = request.args.get('user_id', None)
    
    # Build query
    query = CryptoLog.query
    
    # Apply filters
    if operation_filter != 'all':
        query = query.filter_by(operation=operation_filter)
    
    if success_filter == 'success':
        query = query.filter_by(success=True)
    elif success_filter == 'failure':
        query = query.filter_by(success=False)
    
    if user_filter:
        try:
            user_id = int(user_filter)
            query = query.filter_by(user_id=user_id)
        except ValueError:
            pass
    
    # Order by most recent first
    logs = query.order_by(CryptoLog.timestamp.desc()).limit(100).all()
    
    # Get unique operations for filter dropdown
    all_operations = db.session.query(CryptoLog.operation).distinct().all()
    operations = [op[0] for op in all_operations]
    
    return render_template('crypto_logs.html',
                         logs=logs,
                         operations=operations,
                         operation_filter=operation_filter,
                         success_filter=success_filter,
                         user_filter=user_filter)


@main.route('/profile')
@login_required
def profile():
    """
    View user profile with key information.
    
    Displays user role, public key fingerprint, and key generation timestamp.
    """
    return render_template('profile.html', user=current_user)
